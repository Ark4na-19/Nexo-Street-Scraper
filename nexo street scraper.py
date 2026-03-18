
"""
Web Scraper para nexostreet.co → Excel con imágenes incrustadas
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Requisitos:
    pip install selenium webdriver-manager pandas openpyxl pillow requests
"""

import os, re, time, json, io, requests
from dataclasses import dataclass, asdict
from typing import Optional
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager


# ─────────────────────────────────────────────
# Configuración
# ─────────────────────────────────────────────
BASE_URL        = "https://nexostreet.co"
COLLECTIONS_URL = f"{BASE_URL}/collections/all"
HEADLESS        = False
SCROLL_PAUSE    = 1.5
PAGE_LOAD_WAIT  = 10
IMAGES_DIR      = "imagenes_productos"
OUTPUT_EXCEL    = "nexostreet_catalogo.xlsx"
OUTPUT_JSON     = "nexostreet_productos.json"

# Dimensiones de celda para la imagen en Excel
IMG_WIDTH_PX    = 120   # ancho imagen en Excel (píxeles aprox)
IMG_HEIGHT_PX   = 150   # alto imagen en Excel
ROW_HEIGHT_PT   = 115   # alto de fila en puntos (≈ IMG_HEIGHT_PX)
COL_IMG_WIDTH   = 18    # ancho columna imagen en unidades Excel


# ─────────────────────────────────────────────
# Modelo de datos
# ─────────────────────────────────────────────
@dataclass
class Producto:
    nombre: str
    precio: str
    precio_comparacion: Optional[str]
    url: str
    imagen_local: Optional[str]
    imagen_url: Optional[str]
    disponible: bool
    descripcion: Optional[str] = None


# ─────────────────────────────────────────────
# Driver
# ─────────────────────────────────────────────
def crear_driver():
    opt = Options()
    if HEADLESS:
        opt.add_argument("--headless=new")
    opt.add_argument("--no-sandbox")
    opt.add_argument("--disable-dev-shm-usage")
    opt.add_argument("--window-size=1920,1080")
    opt.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
    )
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opt)


# ─────────────────────────────────────────────
# Descarga de imagen
# ─────────────────────────────────────────────
def nombre_seguro(nombre: str) -> str:
    nombre = re.sub(r'[\\/*?:"<>|]', "", nombre).strip().replace(" ", "_")
    return nombre[:80]


def descargar_imagen(url: str, nombre_producto: str, carpeta: str) -> Optional[str]:
    if not url:
        return None
    url_limpia = url.split("?")[0]
    ext = Path(urlparse(url_limpia).path).suffix or ".jpg"
    ruta = os.path.join(carpeta, nombre_seguro(nombre_producto) + ext)
    if os.path.exists(ruta):
        return ruta
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        r = requests.get(url_limpia, headers=headers, timeout=15)
        r.raise_for_status()
        with open(ruta, "wb") as f:
            f.write(r.content)
        return ruta
    except Exception as e:
        print(f"      [!] Error descargando imagen: {e}")
        return None


def preparar_imagen_excel(ruta: str, ancho: int, alto: int) -> Optional[str]:
    """
    Redimensiona la imagen para que quepa bien en la celda Excel.
    Guarda una versión _thumb junto a la original y devuelve su ruta.
    """
    try:
        thumb_ruta = ruta.replace(".", "_thumb.")
        img = PILImage.open(ruta)
        img.thumbnail((ancho, alto), PILImage.LANCZOS)
        # Convertir a RGB si tiene canal alfa (PNG transparente)
        if img.mode in ("RGBA", "P"):
            bg = PILImage.new("RGB", img.size, (255, 255, 255))
            bg.paste(img, mask=img.split()[3] if img.mode == "RGBA" else None)
            img = bg
        img.save(thumb_ruta, "JPEG", quality=85)
        return thumb_ruta
    except Exception as e:
        print(f"      [!] Error procesando imagen: {e}")
        return ruta   # intentar con la original


# ─────────────────────────────────────────────
# Scroll
# ─────────────────────────────────────────────
def scroll_hasta_el_final(driver):
    h = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE)
        nh = driver.execute_script("return document.body.scrollHeight")
        if nh == h:
            break
        h = nh


# ─────────────────────────────────────────────
# URLs de productos
# ─────────────────────────────────────────────
def obtener_urls_productos(driver) -> list[str]:
    print(f"[→] Cargando: {COLLECTIONS_URL}")
    driver.get(COLLECTIONS_URL)
    try:
        WebDriverWait(driver, PAGE_LOAD_WAIT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/products/']"))
        )
    except TimeoutException:
        print("[!] Sin productos.")
        return []
    scroll_hasta_el_final(driver)
    urls = set()
    for a in driver.find_elements(By.CSS_SELECTOR, "a[href*='/products/']"):
        href = a.get_attribute("href")
        if href and "/products/" in href:
            urls.add(href.split("?")[0])
    print(f"[✓] {len(urls)} productos encontrados.\n")
    return list(urls)


# ─────────────────────────────────────────────
# Disponibilidad
# ─────────────────────────────────────────────
def detectar_disponibilidad(driver) -> bool:
    # 1. Botón "Agregar" activo
    for btn in driver.find_elements(By.CSS_SELECTOR,
        "button[name='add'], button[type='submit'][id*='add'], "
        "button[type='submit'][class*='add-to-cart'], "
        "button[type='submit'][class*='product-form'], "
        "input[type='submit'][name='add']"
    ):
        disabled = btn.get_attribute("disabled")
        aria     = (btn.get_attribute("aria-disabled") or "").lower()
        clase    = (btn.get_attribute("class") or "").lower()
        texto    = btn.text.lower()
        if (disabled is None and aria != "true"
                and "sold" not in clase and "agotado" not in clase
                and "sold out" not in texto and "agotado" not in texto):
            return True

    # 2. Indicadores de agotado visibles
    xpath = (
        "//*[contains(@class,'sold-out') or contains(@class,'soldout') "
        "or contains(@class,'out-of-stock') or "
        "( (self::span or self::p or self::div or self::button) and ("
        "normalize-space(text())='Agotado' or normalize-space(text())='Sold out' "
        "or normalize-space(text())='Sin stock' or normalize-space(text())='Out of stock'"
        ") )]"
    )
    for el in driver.find_elements(By.XPATH, xpath):
        if el.is_displayed():
            return False

    # 3. Todos los options disabled
    opts = driver.find_elements(By.CSS_SELECTOR, "option")
    opts_dis = driver.find_elements(By.CSS_SELECTOR, "option[disabled]")
    if opts and len(opts) == len(opts_dis):
        return False

    return True  # fallback


# ─────────────────────────────────────────────
# Scrapear producto
# ─────────────────────────────────────────────
def scrapear_producto(driver, url: str, carpeta: str) -> Optional[Producto]:
    driver.get(url)
    try:
        WebDriverWait(driver, PAGE_LOAD_WAIT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "h1, .product__title"))
        )
    except TimeoutException:
        print(f"  [!] Timeout: {url}")
        return None

    def txt(sel, default=""):
        try: return driver.find_element(By.CSS_SELECTOR, sel).text.strip()
        except NoSuchElementException: return default

    def attr(sel, a):
        try: return driver.find_element(By.CSS_SELECTOR, sel).get_attribute(a)
        except NoSuchElementException: return None

    nombre = txt(".product__title") or txt("h1.product-single__title") or txt("h1")
    precio = (txt(".price-item--regular") or txt(".price__regular .price-item")
              or txt(".product__price") or txt("[class*='price']"))
    precio_comp = txt(".price-item--compare-at") or txt(".price__was") or None
    if precio_comp == precio:
        precio_comp = None

    # Imagen mayor resolución
    imagen_url = None
    srcset = (attr(".product__media img", "srcset")
              or attr(".product-single__photo img", "srcset")
              or attr("img.product-featured-img", "srcset"))
    if srcset:
        partes = [p.strip() for p in srcset.split(",") if p.strip()]
        if partes:
            imagen_url = partes[-1].split(" ")[0]
    if not imagen_url:
        imagen_url = (attr(".product__media img", "src")
                      or attr(".product-single__photo img", "src")
                      or attr("img.product-featured-img", "src"))
    if imagen_url and imagen_url.startswith("//"):
        imagen_url = "https:" + imagen_url

    imagen_local = descargar_imagen(imagen_url, nombre, carpeta) if imagen_url and nombre else None

    descripcion = (txt(".product__description") or txt(".product-single__description")
                   or txt("[class*='description']") or None)
    if descripcion:
        descripcion = descripcion[:400]

    return Producto(
        nombre=nombre, precio=precio, precio_comparacion=precio_comp,
        url=url, imagen_local=imagen_local, imagen_url=imagen_url,
        disponible=detectar_disponibilidad(driver), descripcion=descripcion,
    )


# ─────────────────────────────────────────────
# Generar Excel con imágenes incrustadas
# ─────────────────────────────────────────────
def generar_excel(productos: list[Producto], ruta: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo Nexo Street"

    # ── Paleta de colores ────────────────────────────────────────────
    COLOR_HEADER_BG  = "1A1A2E"   # azul muy oscuro (marca oscura)
    COLOR_HEADER_FG  = "FFFFFF"
    COLOR_FILA_PAR   = "F5F5F5"
    COLOR_FILA_IMPAR = "FFFFFF"
    COLOR_DISPONIBLE = "27AE60"   # verde
    COLOR_AGOTADO    = "E74C3C"   # rojo
    COLOR_BORDE      = "DDDDDD"

    borde_fino = Border(
        left=Side(style="thin", color=COLOR_BORDE),
        right=Side(style="thin", color=COLOR_BORDE),
        top=Side(style="thin", color=COLOR_BORDE),
        bottom=Side(style="thin", color=COLOR_BORDE),
    )

    # ── Logo / título ────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = "NEXO STREET — Catálogo de Productos"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color=COLOR_HEADER_FG)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Encabezados ──────────────────────────────────────────────────
    headers = ["Imagen", "Nombre", "Precio", "Precio Anterior", "Disponibilidad", "URL", "Descripción"]
    col_widths = [COL_IMG_WIDTH, 30, 16, 16, 16, 40, 50]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=11, color=COLOR_HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde_fino
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[2].height = 28

    # ── Filas de datos ───────────────────────────────────────────────
    for row_idx, p in enumerate(productos, start=3):
        es_par = (row_idx % 2 == 0)
        bg_color = COLOR_FILA_PAR if es_par else COLOR_FILA_IMPAR
        fill = PatternFill("solid", fgColor=bg_color)

        ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT

        # Columna A: imagen (se añade después)
        ws.cell(row=row_idx, column=1).fill = fill
        ws.cell(row=row_idx, column=1).border = borde_fino

        # Columna B: nombre
        c = ws.cell(row=row_idx, column=2, value=p.nombre)
        c.font = Font(name="Arial", bold=True, size=11)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.fill = fill; c.border = borde_fino

        # Columna C: precio
        c = ws.cell(row=row_idx, column=3, value=p.precio)
        c.font = Font(name="Arial", size=11, bold=True, color="1A1A2E")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill; c.border = borde_fino

        # Columna D: precio anterior
        c = ws.cell(row=row_idx, column=4, value=p.precio_comparacion or "—")
        c.font = Font(name="Arial", size=10, strike=bool(p.precio_comparacion), color="888888")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill; c.border = borde_fino

        # Columna E: disponibilidad
        disp_texto = "✓ Disponible" if p.disponible else "✗ Agotado"
        disp_color = COLOR_DISPONIBLE if p.disponible else COLOR_AGOTADO
        c = ws.cell(row=row_idx, column=5, value=disp_texto)
        c.font = Font(name="Arial", bold=True, size=11, color=disp_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill; c.border = borde_fino

        # Columna F: URL (hipervínculo)
        c = ws.cell(row=row_idx, column=6, value=p.url)
        c.hyperlink = p.url
        c.font = Font(name="Arial", size=10, color="0563C1", underline="single")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.fill = fill; c.border = borde_fino

        # Columna G: descripción
        c = ws.cell(row=row_idx, column=7, value=p.descripcion or "")
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.fill = fill; c.border = borde_fino

        # ── Insertar imagen ─────────────────────────────────────────
        if p.imagen_local and os.path.exists(p.imagen_local):
            thumb = preparar_imagen_excel(p.imagen_local, IMG_WIDTH_PX, IMG_HEIGHT_PX)
            if thumb and os.path.exists(thumb):
                try:
                    xl_img = XLImage(thumb)
                    xl_img.width  = IMG_WIDTH_PX
                    xl_img.height = IMG_HEIGHT_PX
                    # Ancla: columna A de la fila actual con pequeño margen
                    celda_ancla = f"A{row_idx}"
                    ws.add_image(xl_img, celda_ancla)
                    print(f"      🖼  Imagen añadida al Excel: {thumb}")
                except Exception as e:
                    print(f"      [!] No se pudo añadir imagen al Excel: {e}")

    # ── Freeze panes (encabezados fijos al hacer scroll) ────────────
    ws.freeze_panes = "B3"

    # ── Hoja de resumen ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Resumen del Catálogo"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_FG)
    ws2["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws2.merge_cells("A1:B1")
    ws2.row_dimensions[1].height = 28
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20

    total = len(productos)
    disponibles = sum(1 for p in productos if p.disponible)
    agotados = total - disponibles

    resumen_data = [
        ("Total de productos", f"=COUNTA(Catálogo!B3:B{total+2})"),
        ("Disponibles",        disponibles),
        ("Agotados",           agotados),
    ]
    for r, (label, val) in enumerate(resumen_data, start=2):
        ws2.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=11)
        ws2.cell(row=r, column=2, value=val).font   = Font(name="Arial", size=11)
        ws2.row_dimensions[r].height = 20

    wb.save(ruta)
    print(f"\n[✓] Excel guardado: {os.path.abspath(ruta)}")


# ─────────────────────────────────────────────
# Guardar JSON
# ─────────────────────────────────────────────
def guardar_json(productos: list[Producto], path: str = OUTPUT_JSON):
    with open(path, "w", encoding="utf-8") as f:
        json.dump([asdict(p) for p in productos], f, ensure_ascii=False, indent=2)
    print(f"[✓] JSON guardado: {path}")


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────
def main():
    os.makedirs(IMAGES_DIR, exist_ok=True)
    print(f"[📁] Imágenes → {os.path.abspath(IMAGES_DIR)}\n")

    driver = crear_driver()
    productos: list[Producto] = []

    try:
        urls = obtener_urls_productos(driver)
        for i, url in enumerate(urls, 1):
            print(f"  [{i}/{len(urls)}] {url}")
            p = scrapear_producto(driver, url, IMAGES_DIR)
            if p:
                productos.append(p)
                disp = "✓ disponible" if p.disponible else "✗ agotado"
                print(f"      {disp} | {p.nombre} — {p.precio}")
            time.sleep(0.8)
    finally:
        driver.quit()

    print(f"\n[✓] Total productos: {len(productos)}")
    guardar_json(productos)
    generar_excel(productos, OUTPUT_EXCEL)

    print("\n── Archivos generados ───────────────────────────────────")
    print(f"  📊 Excel:  {os.path.abspath(OUTPUT_EXCEL)}")
    print(f"  📄 JSON:   {os.path.abspath(OUTPUT_JSON)}")
    print(f"  📁 Imgs:   {os.path.abspath(IMAGES_DIR)}/")


if __name__ == "__main__":
    main()

